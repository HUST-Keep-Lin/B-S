import math
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook


class Running:
    """使用running.csv的数据实现PDR算法， 类的属性均为list"""
    def __init__(self):
        # 记录running提供的数据，以list方式存在属性值当中
        self.accx = []
        self.accy = []
        self.accz = []
        self.gyroscopex = []
        self.gyroscopey = []
        self.gyroscopez = []
        self.timestamp = []
        self.sample_batch = []
        # 记录迈步的时刻，为Z轴加速度取极大值的时刻
        self.step_time = []
        # 记录步长
        self.length = []
        # 记录z轴最大最小加速度
        self.amax = []
        self.amin = []
        # 记录航向角,此为绝对航向角
        self.arg = []
        self.right_arg = []
        # 记录PDR算法得出的位置坐标
        self.position = []

    def peak_fre(self):
        """利用z轴加速度accz使用波峰测频法计算频率, 并记录加速度极大值以及极小值"""
        step = 0
        amax, amin = [], []
        temp_index = 0
        for index, value in enumerate(self.accz):
            if index == 0:
               continue
            if (value - self.accz[index - 1]) > 0:
                while index < len(self.accz) - 1 and self.accz[index] < self.accz[index + 1]:
                    index += 1
                # 在波峰的前bias个值中找波谷, 由表中数据可得两个波峰大致上6个点，所以可以取bias为3或者4
                # 同时，index得到的最大值也有可能是伪最大值，需要进行进一步判断和处理
                bias = 4
                if bias < index:
                    temp = self.accz[index - bias: index]
                    self.accz = np.array(self.accz)
                    i = np.where(self.accz == min(temp))[0][0]
                    # 有可能出现误判的信息，需要进行剔除
                    if temp_index >= i or [self.timestamp[i], self.accz[i]] in amin:

                        last = amax.pop()
                        if self.accz[index] > last[1]:
                            amax.append([self.timestamp[index], self.accz[index]])
                            temp_index = index
                        else:
                            amax.append(last)
                    else:
                        step += 1
                        temp_index = index
                        print(index)
                        amax.append([self.timestamp[index], self.accz[index]])
                        amin.append([self.timestamp[i], self.accz[i]])


        for value in amax:
            self.step_time.append(value[0])

        self.amax = amax
        self.amin = amin

    def step_length(self):
        """使用Weinberg算法估算步长, 取K = 0.7"""
        k = 0.7
        for i in range(len(self.amax)):
            self.length.append(k * pow(self.amax[i][1] - self.amin[i][1], 0.25))


    def get_arg(self, bias):
        """其中bias为初始的航向角度"""
        length = len(self.timestamp)
        alpha_new = bias
        # angle存储每一个点对应的角度值
        angle = []
        for i in range(length - 1):

            t = (self.timestamp[i + 1] - self.timestamp[i]) / 1000

            m = np.sqrt(self.accx[i] ** 2 + self.accy[i] ** 2 + (self.accz[i] - 1) ** 2)
            change = 0.08 * (
                        self.accx[i] * self.gyroscopex[i] + self.accy[i] * self.gyroscopey[i] + (self.accz[i] - 1) *
                        self.gyroscopez[i]) * t / m

            # 数据修正, 滤波器实现光滑
            if i <= 2 * length / 5 or i >= 3 * length / 5:
                change = change / 1.2
            else:
                change = change * 1.6

            alpha_new += change
            alpha_new = round(alpha_new, 2)
            angle.append(alpha_new)

        self.arg = angle

    def get_position(self, x, y):
        self.position.append([x, y])
        index = 0
        bis = 5
        for i in range(len(self.timestamp)):
            if self.timestamp[i] in self.step_time:
                # 利用PDR算法计算每一步的位移，并进行存储
                x += self.length[index] * math.cos(self.arg[i] * np.pi / 180)
                y += self.length[index] * math.sin(self.arg[i] * np.pi / 180)
                self.right_arg.append(self.arg[i])
                # 进行数据处理，位置坐标保留两位小数
                x = round(x, 2)
                y = round(y, 2)
                self.position.append([x, y])
                # print(self.arg[i])
                index += 1
                # print(x, y)
        return self.position


def get_error(x_pre, y_pre, x_tru, y_tru):
    """计算误差"""
    # errors保存每一个定位点的误差, cdf保存cdf误差, pre_errors保存百分位误差
    errors = []
    cdf = []
    pre_errors = []
    error50 = []
    error75 = []
    error90 = []
    num = len(x_pre)
    for i in range(num):
        errors.append(np.sqrt((x_pre[i] - x_tru[i]) ** 2 + (y_tru[i] - y_pre[i]) ** 2))

    ave_error = sum(errors) / len(errors)

    for i in np.linspace(min(errors) - 0.5, max(errors) + 0.5, 500):
        cnt = 0
        for error in errors:
            if error < i:
                cnt += 1
        cdf.append([i, cnt/num])
        if round(cnt/num, 1) == 0.5:
            error50.append(i)
        if round(cnt/num, 1) == 0.7:
            error75.append(i)
        if round(cnt/num, 1) == 0.9:
            error90.append(i)

    pre_errors.append(sum(error50)/len(error50))
    pre_errors.append(sum(error75)/len(error75))
    pre_errors.append(sum(error90)/len(error90))

    return errors, ave_error, cdf, pre_errors

def data_proce(x_pre, y_pre, x_tru, y_tru):
    """数据处理，将预测点与真实点变为长度相等"""
    x_pre = x_pre[::2]
    y_pre = y_pre[::2]
    if len(x_pre) >= len(x_tru):
        x_pre = x_pre[:len(x_tru)]
        y_pre = y_pre[:len(y_tru)]
    else:
        x_tru = x_tru[:len(x_pre)]
        y_tru = y_tru[:len(y_pre)]

    return x_pre, y_pre, x_tru, y_tru

if __name__ == '__main__':

    # 读取running的数据,并存储在类running中
    path = 'E:\\Homework\\project\\backend\\src\\main\\resources\\file\\running.csv'
    df = pd.read_csv(path)
    running = Running()
    running.accx = list(df.accx)
    running.accy = list(df.accy)
    running.accz = list(df.accz)
    running.accx = [x / 16384 + 1 for x in running.accx]
    running.accy = [x / 16384 + 1 for x in running.accy]
    running.accz = [x / 16384 + 1 for x in running.accz]
    running.gyroscopex = list(df.gyroscopex)
    running.gyroscopey = list(df.gyroscopey)
    running.gyroscopez = list(df.gyroscopez)
    running.timestamp = list(df.timestamp)

    #print(running.accx)
    # 调用PDR算法，得到步频步长航向
    running.peak_fre()
    running.step_length()
    running.get_arg(-90)

    # 读取position.csv的内容，将其保存在x_pos, y_pos中
    path = 'E:\\Homework\\project\\backend\\src\\main\\resources\\file\\position.csv'
    df = pd.read_csv(path)
    x_pos = list(df.x)
    y_pos = list(df.y)

    # 以position的初始点作为pdr初始位置
    path = 'E:\\Homework\\project\\backend\\src\\main\\resources\\file\\ground_truth.csv'
    df = pd.read_csv(path)
    x_tru = list(df.x)
    y_tru = list(df.y)
    x_begin = x_tru[0]
    y_begin = y_tru[0]
    print("yes")
    print(running.amax)
    print(running.right_arg)
    print(running.step_time)
    # 根据PDR算法进行定位，position是一个嵌套list, 每个元素list内部为每个定位点的坐标
    position = running.get_position(x_begin, y_begin)
    print("oh")
    # x_run, y_run保存预测的位置
    x_run = []
    y_run = []
    for pos in position:
        x_run.append(pos[0])
        y_run.append(pos[1])
    # 对定位点做修正， 由于两种方式计算的步长的区别，需要做改进
#     x_tru = [-1, -1, -1, -1, -1, -1, -0.6, 0.2, 1.2, 1.5, 1.5, 1.5, 1.5, 1.5, 1.5]
#     y_tru = [3.4, 2.2, 1, -0.2, -1.4, -2.6, -3.2, -3.2, -3.2, -2.6, -1.4, -0.2, 1, 2.2, 3.4]

    x_run, y_run, x_tru, y_tru = data_proce(x_run, y_run, x_tru, y_tru)

    # errors是每个点的定位误差，ave_error为平均误差，cdf为cdf误差的点集， pre_errors为百分比误差
    errors, are_error, cdf, pre_errors = get_error(x_run, y_run, x_tru, y_tru)

    cdf_x = []
    cdf_y = []
    for value in cdf:
        cdf_x.append(value[0])
        cdf_y.append(value[1])

    # 绘图
    # plt.plot(cdf_x, cdf_y)
    # plt.xlabel('distance estimation error(m)')
    # plt.ylabel('CDF')
    # plt.title('Cumulative Distribution Function')
    # plt.show()
    #
    # plt.plot(x_run, y_run, label='running')
    # plt.plot(x_pos, y_pos, label='position')
    # plt.plot(x_tru, y_tru, label='ground_truth')
    # plt.xlabel('x')
    # plt.ylabel('y')
    # plt.grid(linestyle="--")
    # plt.legend()
    #
    # plt.show()

    running.right_arg = running.right_arg[::2]
    if len(running.right_arg) > len(x_run):
        running.right_arg = running.right_arg[:len(x_run)]
    else:
        running.right_arg = running.right_arg + [0] * (len(x_run) - len(running.right_arg))

    max_num = len(cdf_x)
    x_run = x_run + [-500] * (len(cdf_x) - len(x_run))
    y_run = y_run + [-500] * (len(cdf_x) - len(y_run))
    running.right_arg = running.right_arg + [-500] * (len(cdf_x) - len(running.right_arg))
    errors = errors + [-500] * (len(cdf_x) - len(errors))
    pre_errors = pre_errors + [-500] * (len(cdf_x) - len(pre_errors))



    # 将PDR数据写入result.xlsx文件
    # 创建数据
    excel_path = "E:\\Homework\\project\\backend\\src\\main\\resources\\file\\infomation.xlsx"  # 表格名称

    wb = Workbook(excel_path)  # 创建xlsx表格
    wb.save(excel_path)  # 保存

    wb = load_workbook(excel_path)  # 导入表格，用于下面进行操作
    wb.create_sheet("sheet1")  # 创建表名
    ws = wb.active  # 激活

    ws.cell(1, 1).value = 'x'
    ws.cell(1, 2).value = 'y'
    ws.cell(1, 3).value = 'arg'
    ws.cell(1, 4).value = 'error'
    ws.cell(1, 5).value = 'cdf_x'
    ws.cell(1, 6).value = 'cdf_y'
    ws.cell(1, 7).value = 'pre_error'
    for row in range(1, len(x_run) + 1):
        ws.cell(row + 1, 1).value = x_run[row - 1]
    for row in range(1, len(y_run) + 1):
        ws.cell(row + 1, 2).value = y_run[row - 1]
    for row in range(1, len(running.right_arg) + 1):
        ws.cell(row + 1, 3).value = running.right_arg[row - 1]
    for row in range(1, len(errors) + 1):
        ws.cell(row + 1, 4).value = errors[row - 1]
    for row in range(1, len(cdf_x) + 1):
        ws.cell(row + 1, 5).value = cdf_x[row - 1]
    for row in range(1, len(cdf_y) + 1):
        ws.cell(row + 1, 6).value = cdf_y[row - 1]
    for row in range(1, len(pre_errors) + 1):
        ws.cell(row + 1, 7).value = pre_errors[row - 1]
    # for row in range(1, len(x_run) + 1):  # 控制行
    #     for column in range(1, len(data_list) + 1):  # 控制列
    #         ws.cell(row + 1, column).value = data_list[column - 1][row - 1]  # 把数据传输到表格中

    wb.save(excel_path)  # 保存表格
    wb.close()  # 关闭

    # print("%s     保存成功！" % excel_path)


